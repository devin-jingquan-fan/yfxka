#!/usr/bin/env python
# _*_ UTF-8 _*_
'''
@project:GBDT+LR-Demo
@author:xiaofei
'''
from openpyxl.workbook import Workbook
import openpyxl
import pandas as pd
import numpy
import xlrd


if __name__ == '__main__':

    print ('读取成功')

#        读取文件
    workbook1=openpyxl.load_workbook("/Users/fanjingquan/Desktop/test1.xlsx")  #文件路径
#        读取第一个工作表
    worksheet=workbook1.worksheets[0]
    

# =============================================================================
# #    读取某块数值，查找替换功能
#     for i in range(1, 10000):
#         for j in range(1, 10):
#             if worksheet.cell(row=i, column=j).value == 'a':
#                 worksheet.cell(row=i, column=j).value = 'b'
# #            print(worksheet.cell(row=i, column=j).value,end=" ")
# #        print()
# =============================================================================
    
    
#        读取第一列的值查找替换
    for x in range(1,50):
        if worksheet.cell(row=x, column=2).value == 'like':
            worksheet.cell(row=x, column=2).value = '=sum(1:1)'
        print(worksheet.cell(row=x, column=2).value,end=" ")
        print()
        
    
# =============================================================================
#     for i in range(1,146):
#         te=0   #一定要先初始化为一个int型
#         te=str(worksheet.cell(row=i,column=1).value) #强制类型转化 不初始化不能完成赋值！
#         worksheet.cell(row=i,column=2).value=worksheet2.cell(row=te,column=2).value
# =============================================================================

#另存为文件
    workbook1.save(filename='/Users/fanjingquan/Desktop/test2.xlsx')